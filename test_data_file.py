# import datetime
# import pathlib
# FILE_PATTERN = "*.txt"
# ARCHIVE = "archive"
# if __name__ == '__main__':
#     date_string = datetime.date.today().strftime("%Y-%m-%d")
#     cur_path = pathlib.Path(".")
#     new_path = cur_path.joinpath(ARCHIVE, date_string)
#     new_path.mkdir()
#     paths = cur_path.glob(FILE_PATTERN)
#     for path in paths:
#         path.rename(new_path.joinpath(path.name))

open('test.txt', 'wb').write(bytes([65, 66, 67, 255, 192,193]))

x = open('test.txt').read()
print(x)
x=open('test.txt', errors='ignore').read()
print(x)
x=open('test.txt', errors='replace').read()
print(x)
x=open('test.txt', errors='surrogateescape').read()
print(x)
x=open('test.txt', errors='backslashreplace').read()
print(x)

import csv
results = [fields for fields in csv.reader(open("temp_data_pipes_00a.txt", newline=''), delimiter="|")]
results = [x for x in range(7)]
print(results)

results = [fields for fields in csv.DictReader(open("temp_data_01.csv",newline=''))]
print(results)

print('_______')
from openpyxl import load_workbook
wb = load_workbook('temp_data_01.xlsx')
results = []
ws = wb.worksheets[0]
for row in ws.iter_rows():
    results.append([cell.value for cell in row])
print(results)

results = [fields for fields in csv.reader(open("temp_data_01.csv",newline=''))]
data_rows = [fields for fields in csv.reader(open("temp_data_01.csv",newline=''))]
# for fields in csv.reader(open("temp_data_01.csv",newline='')):
    # row = []
    # for field in fields:
    #     if field=='Missing':
    #         field = None
    #     elif field.endswith('%'):
    #         field = field[:-1]
    #     # field = field.strip('%')
    #     row.append(field)
    # print(row)
# del data_rows[0]
clean_field = [float(x[13]) for x in data_rows[1:] if x[13] != 'Missing']
average = sum(clean_field) / len(clean_field)
print(average)
coverage_values = [float (x [-1] .strip ("%")) / 100 for x in data_rows[1:]]
print(coverage_values)